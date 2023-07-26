import sys,os   #Demo7_16.py
from PyQt5.QtWidgets import QApplication, QWidget,QMessageBox
from PyQt5.QtCore import pyqtSlot,pyqtSignal
from openpyxl import Workbook,load_workbook

import student  #导入student.py文件

class QmyWidget(QWidget):
    numberSignal = pyqtSignal(int)  #定义信号
    def __init__(self,parent = None):
        super().__init__(parent)
        self.ui = student.Ui_Form()
        self.ui.setupUi(self)
        self.__student = dict()  #记录学生姓名、学号、成绩的字典，关键字是学号
        self.numberSignal.connect(self.isNumberExisting)  #信号与槽函数的连接
    @pyqtSlot()
    def on_btnCalculate_clicked(self):
        s = self.ui.chinese.value()+self.ui.math.value()+self.ui.english.value()
        self.ui.total.setText(str(s))
        template = "{:.1f}".format(s/3)
        self.ui.average.setText(template)

        temp = list()
        temp.append(self.ui.name.text())
        temp.append(self.ui.number.value())
        temp.append(self.ui.chinese.value())
        temp.append(self.ui.math.value())
        temp.append(self.ui.english.value())
        temp.append(s)
        temp.append(float(template))
        self.__student[self.ui.number.value()] = temp

    @pyqtSlot()
    def on_btnSave_clicked(self):
        path = "./student.xlsx"
        if os.path.exists(path):
            wbook = load_workbook(path)
        else:
            wbook = Workbook()
        wsheet = wbook.active
        wsheet.append(["姓名", "学号", "语文", "数学", "英语", "总成绩", "平均分"])
        student_number = self.__student.keys()
        student_number = list(student_number)
        student_number.sort()
        for i in student_number:
            wsheet.append(self.__student[i])
        wbook.save(path)


    def isNumberExisting(self,value):
        if value in self.__student:  #如果学号已经存在
            existing = QMessageBox.question(self,"确认信息","该学号已经存在，是否覆盖？",
                            QMessageBox.Yes | QMessageBox.No)   #提示对话框
            if existing == QMessageBox.No:  #如果不覆盖，需要重新输入学号
                self.ui.number.setValue(0)  # 学号设置为0，等待重新输入
    def on_number_editingFinished(self):  #输入学号完成时的槽函数（自动关联的槽函数）
        self.numberSignal.emit(self.ui.number.value())  #发射信号，信号参数是学号

if  __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = QmyWidget()
    myWindow.show()
    n = app.exec()
    sys.exit(n)
